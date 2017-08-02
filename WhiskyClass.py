from os import path
from openpyxl import load_workbook
import pandas as pd
import proforma
from WhiskyGroupClass import WhiskyGroup


class Whisky(WhiskyGroup):
    # A class for creating objects for individual whiskies.
    def __init__(self, distillery, barrel_code, bond_year=False, bond_quarter=False, fill_period=False):
        WhiskyGroup.__init__(self, distillery, barrel_code)
        self.category_name = ''
        if fill_period == False:
            # Fill period was not given, instead the bond year and quarter was given.
            self.bond_year = str(bond_year)  # Year it was barreled YYYY
            self.bond_quarter = bond_quarter
        else:
            # Fill period was given
            self.__deconstruct_fill_period(fill_period)

        self.filename = self.distillery + "_" + str(self.bond_year) + self.bond_quarter + "_" + self.barrel_code + "_prices.xlsx"

        self.historic_data_dir = path.join(self.historic_group_directory, self.filename)
        self.current_data_dir = path.join(self.current_group_directory, self.filename)

        self.whisky_code = self.distillery + self.bond_year + self.bond_quarter + self.barrel_code

        self.start_date = self.get_start_date()


    def get_worksheet(self, dataset):
        if dataset == 'historic' or dataset == 'historical':
            directory = self.historic_data_dir
        elif dataset == 'current':
            directory = self.current_data_dir
        else:
            dataset_retry = input("Please enter either \"historic\" or \"current\".")
            self.get_worksheet(dataset_retry)

        wb = load_workbook(directory)
        ws = wb.get_active_sheet()
        return ws

    def get_historic_df(self):
        pass

    def get_current_df(self):
        # Reads the workbook into a formatted data frame.
        df = pd.read_excel(self.current_data_dir, columns=proforma.STANDARD_COLUMNS)
        df.set_index("Date", inplace=True)
        return df

    def __deconstruct_fill_period(self, fill_period):
        if fill_period[4] == '/':
            # fill_period looks like: YYYY/QX
            separated_string = fill_period.replace('/', ' ')
            self.bond_year = separated_string.split()[0]
            self.bond_quarter = separated_string.split()[1]
        else:
            # fill_period looks like: YYYYQ4
            separated_string = fill_period.replace('Q', ' Q')  # Put a space before the Q
            self.bond_year = separated_string.split()[0]
            self.bond_quarter = separated_string.split()[1]

    def get_start_date(self):
        wb = load_workbook(proforma.METADATA_FILENAME)
        ws = wb.get_active_sheet()
        row_index = '2'  # Starting at row 2 which is the first row below the column title.
        code_col = ws["F"][1:]
        for cell in code_col:
            if self.whisky_code == cell.value:
                # We have found the row corresponding to this whisky's entry in the metadata file.
                whisky_start_date = ws["E" + row_index].value
                if whisky_start_date is None:
                    print("This whisky has no start date recorded.")
                    return None
                else:
                    return whisky_start_date
            row_index = str(int(row_index) + 1)
        return None  # The whisky doesn't exist in the metadata file.
