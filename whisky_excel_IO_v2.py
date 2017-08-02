'''
This should be run first.
It scrapes for the price data of each whisky to the current date and saves the data in excel files

v2
ScrapeNewData now uses the Whisky class.
'''

from whisky_scraper import GetOldData, GetNewData
from openpyxl import load_workbook, Workbook
from datetime import timedelta
from os import listdir, getcwd, path, mkdir
import pandas as pd
from WhiskyClass import Whisky
from proforma import STANDARD_COLUMNS, METADATA_FILENAME, SAVE_PATH_CURRENT, SAVE_PATH_HISTORICAL

URL_BASE = "https://www.whiskyinvestdirect.com/"

'''
This 
The start dates for each whisky has been input manually.

FUTURE IMPROVEMENTS:
    Make it so that scraped data is added to the already existing excel file, so that if they remove older data,
    we still have a copy of it.
    Backup.
    Only whiskies that are in the whisky_metadata_with_dates.xlsx file are scraped from their respective web chart.
        This is an issue because the list of whiskies on the site changes through time.
        We have a script to collate a list of all whiskies (whisky_metadata_excel), but we need to manually input the
        start date for each whisky.
        SOLUTION: Use the build the whisky_metadata_excel script to simply add new whiskies to the file without changing
        the others and then notify the user to manually input the start date in the file.
'''
def scrape_old_data():
    metadata_wb = load_workbook(METADATA_FILENAME)  # Made with whisky_metadata_excel.py and dates manually
    metadata_ws = metadata_wb.active


    distilleries = metadata_ws["A2":"A66"]  # Make this dynamic later
    fillPeriods = metadata_ws["B2":"B66"]
    barrelCodes = metadata_ws["C2":"C66"]
    startDates = metadata_ws["E2":"E66"]  # The start dates has been manually entered.


    for index in range(0, len(distilleries)):
        distillery = distilleries[index][0].value  # For some reason the cell is in a tuple with it being the only element
        fill_period = fillPeriods[index][0].value
        barrel_code = barrelCodes[index][0].value
        start_date = startDates[index][0].value # This is already in datetime format for some reason

        whisky = Whisky(distillery, barrel_code, fill_period=fill_period)

        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws["A1"] = "Date"
        output_ws["B1"] = fill_period + " Price"
        # output_ws["C1"] = "Percent Change"
        create_directory(whisky, "historic")
        dest = whisky.historic_data_dir


        url = URL_BASE + distillery + "/" + fill_period + "/" + barrel_code + "/chart.do"
        print(url)

        price_data_json = GetOldData(url)

        initial_date = start_date
        first_datum = True
        row_count = 2
        for trade_entry in price_data_json:
            current_day = trade_entry["day"]
            price = trade_entry["priceAvg"]

            if first_datum:
                first_datum = False
                old_day = current_day
                current_date = initial_date
                # old_price = price
                # first_price = price

                output_ws["A" + str(row_count)] = current_date
                output_ws["B" + str(row_count)] = price

            else:
                date_increment = current_day - old_day
                old_day = current_day
                current_date = initial_date + timedelta(days=date_increment)
                initial_date = current_date

                output_ws["A" + str(row_count)] = current_date
                output_ws["B" + str(row_count)] = price

                ''' Uncomment to add in a percent increase column '''
                # pctChange = (price/oldPrice) - 1
                # output_ws["C" + str(rowCount)].number_format = '0.00%'
                # output_ws["C" + str(rowCount)] = pctChange
                # oldPrice = price
            row_count += 1

        # output_ws["B" + str(rowCount)] = "TOTAL"
        # overallChange = (price/firstPrice) - 1
        # output_ws["C" + str(rowCount)].number_format = '0.00%'
        # output_ws["C" + str(rowCount)] = overallChange

        output_wb.save(filename=dest)


'''
This updates prices for all whiskies once and saves it in their respective .xlsx files. The files are separate
from the files containing historic data.
'''
def scrape_new_data():
    currentMarketJSON = GetNewData()

    whiskyList = currentMarketJSON["market"]["pitches"]
    for whisky_file in whiskyList:
        whisky = Whisky(whisky_file["distillery"], whisky_file["barrelTypeCode"], whisky_file["bondYear"],
                        whisky_file["bondQuarter"])

        category_name = whisky_file["categoryName"]

        filename = whisky.filename
        group_directory = whisky.current_group_directory
        file_directory = whisky.current_data_dir

        file_is_new = True
        try:
            directory_list = listdir(group_directory)
            if filename in directory_list:
                df = whisky.get_current_df()
                file_is_new = False
        except:
            print("Directory does not exist; will be created.")
            create_directory(whisky, "current")

        new_row = extract_data_from_json(whisky_file, currentMarketJSON)

        temp_df = pd.DataFrame([new_row], columns=STANDARD_COLUMNS)
        temp_df.set_index("Date", inplace=True)
        if file_is_new:
            df = temp_df
        else:
            df = df.append(temp_df)

        df.to_excel(file_directory, sheet_name="Price and Volume")


def create_directory(whisky_object, dataset):
    # Creates the distillery directory and group directory for either historic or current dataset.
    print("Directory does not exist; will be created.")
    if dataset == "historic":
        if whisky_object.distillery not in listdir(SAVE_PATH_HISTORICAL):
            # Directory for the distillery does not exist.
            mkdir(whisky_object.historic_distillery_directory)
        working_directory = path.join(whisky_object.historic_distillery_directory)
        if whisky_object.barrel_code not in listdir(working_directory):
            # Directory for the barrel code within the distillery directory does not exist.
            mkdir(whisky_object.historic_group_directory)

    elif dataset == "current":
        if whisky_object.distillery not in listdir(SAVE_PATH_CURRENT):
            mkdir(whisky_object.current_distillery_directory)
        working_directory = path.join(whisky_object.current_distillery_directory)
        if whisky_object.barrel_code not in listdir(working_directory):
            mkdir(whisky_object.current_group_directory)

'''
This extracts the price and volume data for buy/sell from the passed json. 
It returns a single row containing lowest/middle/highest buy & sell prices and their respective quantities.
'''
def extract_data_from_json(whisky_file, currentMarketJSON):
    buyPriceOrderedList = []
    buyQuantityOrderedList = []
    sellPriceOrderedList = []
    sellQuantityOrderedList = []
    totalBuyQuantity = averageBuyPrice = 0
    totalSellQuantity = averageSellPrice = 0

    for priceTier in range(0, 3):
        try:
            buyPriceOrderedList.append(whisky_file["sellPrices"][priceTier]["limit"])
            buyQuantityOrderedList.append(whisky_file["sellPrices"][priceTier]["quantity"])
            totalBuyQuantity += buyQuantityOrderedList[priceTier]
            averageBuyPrice += buyPriceOrderedList[priceTier] * buyQuantityOrderedList[priceTier]
        except:
            buyPriceOrderedList.append("N/A")
            buyQuantityOrderedList.append("N/A")

        try:
            sellPriceOrderedList.append(whisky_file["buyPrices"][priceTier]["limit"])
            sellQuantityOrderedList.append(whisky_file["buyPrices"][priceTier]["quantity"])
            totalSellQuantity += sellQuantityOrderedList[priceTier]
            averageSellPrice += sellPriceOrderedList[priceTier] * sellQuantityOrderedList[priceTier]
        except:
            sellPriceOrderedList.append("N/A")
            sellQuantityOrderedList.append("N/A")

    try:
        averageBuyPrice = averageBuyPrice / totalBuyQuantity
    except:
        averageBuyPrice = "N/A"
    try:
        averageSellPrice = averageSellPrice / totalSellQuantity
    except:
        averageSellPrice = "N/A"

    time = currentMarketJSON["updateTimeString"]

    # This puts everything into a nice single row so that it can be added to the df.
    row = [time,
           buyPriceOrderedList[0], buyQuantityOrderedList[0],
           buyPriceOrderedList[1], buyQuantityOrderedList[1],
           buyPriceOrderedList[2], buyQuantityOrderedList[2],
           averageBuyPrice, totalBuyQuantity,
           sellPriceOrderedList[0], sellQuantityOrderedList[0],
           sellPriceOrderedList[1], sellQuantityOrderedList[1],
           sellPriceOrderedList[2], sellQuantityOrderedList[2],
           averageSellPrice, totalSellQuantity
           ]

    return row
