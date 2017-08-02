'''
This should be run first.
It scrapes for the price data of each whisky to the current date and saves the data in excel files
'''

from whisky_scraper import GetOldData, GetNewData
from openpyxl import load_workbook, Workbook
from datetime import timedelta
from os import listdir, getcwd, path
import pandas as pd

SAVE_PATH_HISTORICAL = path.join(getcwd(), path.join("spreadsheets", "historical"))
SAVE_PATH_CURRENT = path.join(getcwd(), path.join("spreadsheets", "current"))
URL_BASE = "https://www.whiskyinvestdirect.com/"

# STANDARD_COLUMNS = ["Date",
#                     "Cheapest Buy Price", "Cheapest Buy Quantity",
#                     "Middle Buy Price", "Middle Buy Quantity",
#                     "Expensive Buy Price", "Expensive Buy Quantity",
#                     "Average Buy Price", "Average Buy Quantity"]

STANDARD_COLUMNS = ["Date",
                    "Cheapest Buy Price", "Cheapest Buy Quantity",
                    "Middle Buy Price", "Middle Buy Quantity",
                    "Expensive Buy Price", "Expensive Buy Quantity",
                    "Average Buy Price", "Average Buy Quantity",
                    "Cheapest Sell Price", "Cheapest Sell Quantity",
                    "Middle Sell Price", "Middle Sell Quantity",
                    "Expensive Sell Price", "Expensive Sell Quantity",
                    "Average Sell Price", "Average Sell Quantity"
                    ]
'''
This 
The start dates for each whisky has been input manually.
'''
def ScrapeOldData():
    metadata_wb = load_workbook("whisky_metadata_with_dates.xlsx")  # Made with whisky_metadata_excel.py and dates manually
    metadata_ws = metadata_wb.active


    distilleries = metadata_ws["A2":"A66"]  # Make this dynamic later
    fillPeriods = metadata_ws["B2":"B66"]
    barrelCodes = metadata_ws["C2":"C66"]
    startDates = metadata_ws["E2":"E66"]  # The start date has been manually entered.


    for index in range(0, 65):
        distillery = distilleries[index][0].value  # For some reason the cell is in a tuple with it being the only element
        fillPeriod = fillPeriods[index][0].value
        barrelCode = barrelCodes[index][0].value
        startDate = startDates[index][0].value # This is already in datetime format for some reason

        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws["A1"] = "Date"
        output_ws["B1"] = fillPeriod + " Price"
        # output_ws["C1"] = "Percent Change"
        dest_filename = distillery + "_" + fillPeriod.replace("/", "") + "_" + barrelCode + "_prices.xlsx"
        dest = path.join(SAVE_PATH_HISTORICAL, dest_filename)  # To save files in a different folder


        url = URL_BASE + distillery + "/" + fillPeriod + "/" + barrelCode + "/chart.do"
        print(url)

        priceDataJSON = GetOldData(url) 

        initialDate = startDate
        firstDatum = True
        rowCount = 2
        for tradeEntry in priceDataJSON:
            currentDay = tradeEntry["day"]
            price = tradeEntry["priceAvg"]

            if firstDatum:
                firstDatum = False
                oldDay = currentDay
                currentDate = initialDate
                # oldPrice = price
                # firstPrice = price

                output_ws["A" + str(rowCount)] = currentDate
                output_ws["B" + str(rowCount)] = price

            else:
                dateIncrement = currentDay - oldDay
                oldDay = currentDay
                currentDate = initialDate + timedelta(days = dateIncrement)
                initialDate = currentDate

                output_ws["A" + str(rowCount)] = currentDate
                output_ws["B" + str(rowCount)] = price

                ''' Uncomment to add in a percent increase column '''
                # pctChange = (price/oldPrice) - 1
                # output_ws["C" + str(rowCount)].number_format = '0.00%'
                # output_ws["C" + str(rowCount)] = pctChange
                # oldPrice = price
            rowCount += 1

        # output_ws["B" + str(rowCount)] = "TOTAL"
        # overallChange = (price/firstPrice) - 1
        # output_ws["C" + str(rowCount)].number_format = '0.00%'
        # output_ws["C" + str(rowCount)] = overallChange

        output_wb.save(filename = dest)


'''
This updates prices for all whiskies once and saves it in their respective .xlsx files. The files are separate
from the files containing historic data.
'''
def ScrapeNewData():
    directoryList = listdir(SAVE_PATH_CURRENT)
    currentMarketJSON = GetNewData()

    whiskyList = currentMarketJSON["market"]["pitches"]
    for whisky in whiskyList:
        distillery = whisky["distillery"]
        categoryName = whisky["categoryName"]
        barrelTypeCode = whisky["barrelTypeCode"]
        bondYear = whisky["bondYear"]
        bondQuarter = whisky["bondQuarter"]

        filename = distillery + "_" + str(bondYear) + bondQuarter + "_" + barrelTypeCode + "_prices.xlsx"
        file = path.join(SAVE_PATH_CURRENT, filename)
        if filename in directoryList:
            df = pd.read_excel(file, columns=STANDARD_COLUMNS)
            df.set_index("Date", inplace=True)
            fileIsNew = False
        else:
            df = pd.DataFrame(columns=["Date", "Price", "Quantity"])
            fileIsNew = True


        buyPriceOrderedList = []
        buyQuantityOrderedList = []
        sellPriceOrderedList = []
        sellQuantityOrderedList = []
        totalBuyQuantity = averageBuyPrice = 0
        totalSellQuantity = averageSellPrice = 0


        for priceTier in range(0,3):
            try:
                buyPriceOrderedList.append(whisky["sellPrices"][priceTier]["limit"])
                buyQuantityOrderedList.append(whisky["sellPrices"][priceTier]["quantity"])
                totalBuyQuantity += buyQuantityOrderedList[priceTier]
                averageBuyPrice += buyPriceOrderedList[priceTier] * buyQuantityOrderedList[priceTier]
            except:
                buyPriceOrderedList.append("N/A")
                buyQuantityOrderedList.append("N/A")

            try:
                sellPriceOrderedList.append(whisky["buyPrices"][priceTier]["limit"])
                sellQuantityOrderedList.append(whisky["buyPrices"][priceTier]["quantity"])
                totalSellQuantity += sellQuantityOrderedList[priceTier]
                averageSellPrice += sellPriceOrderedList[priceTier] * sellQuantityOrderedList[priceTier]
            except:
                sellPriceOrderedList.append("N/A")
                sellQuantityOrderedList.append("N/A")

        try:
            averageBuyPrice = averageBuyPrice / totalBuyQuantity
        except:
            averageBuyPrice = "N/A"
        try:
            averageSellPrice = averageSellPrice / totalSellQuantity
        except:
            averageSellPrice = "N/A"


        time = currentMarketJSON["updateTimeString"]

        row = [time,
               buyPriceOrderedList[0], buyQuantityOrderedList[0],
               buyPriceOrderedList[1], buyQuantityOrderedList[1],
               buyPriceOrderedList[2], buyQuantityOrderedList[2],
               averageBuyPrice, totalBuyQuantity,
               sellPriceOrderedList[0], sellQuantityOrderedList[0],
               sellPriceOrderedList[1], sellQuantityOrderedList[1],
               sellPriceOrderedList[2], sellQuantityOrderedList[2],
               averageSellPrice, totalSellQuantity
               ]

        temp_df = pd.DataFrame([row], columns=STANDARD_COLUMNS)
        temp_df.set_index("Date", inplace=True)
        if fileIsNew:
            df = temp_df
        else:
            df = df.append(temp_df)

        df.to_excel(file, "Price and Volume")
    print(file)