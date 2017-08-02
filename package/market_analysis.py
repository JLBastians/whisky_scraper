from openpyxl import load_workbook
from whisky_excel_IO_v2 import SAVE_PATH_CURRENT
from os import listdir
import os.path
from WhiskyClass import Whisky
import datetime
import pandas as pd
from WhiskyGroupClass import WhiskyGroup
import matplotlib.pyplot as plt

AVERAGE_PRICE_COLUMNS = {"Buy": "H", "Sell": "P"}  # Column index in excel.
LENGTH_OF_QUARTER = 90  # Days
DATE_COLUMN_LABEL = "Date"

'''
Imports data from the "current" .xlsx files for each whisky in the directory.
Calculates the difference between the highest and lowest prices for buying/selling over an arbitrary passed period.
E.G: Volatility(30, "minutes") calculates the difference between the highs/lows over a 30 minute period. [POORLY DEFINED]
'''
def Volatility():
    directoryList = listdir(SAVE_PATH_CURRENT)
    directoryList.remove("backup") # Remove the backup folder so that only the .xlsx files outside it remain.

    for file in directoryList:
        fullFileName = os.path.join(SAVE_PATH_CURRENT, file)
        wb = load_workbook(fullFileName)
        ws = wb.get_active_sheet()
        avgBuyCol = ws[AVERAGE_PRICE_COLUMNS["Buy"]]
        avgSellCol = ws[AVERAGE_PRICE_COLUMNS["Sell"]]

        whiskyNameData = (file.replace("_", " ")).split()
        if len(whiskyNameData) > 4:
            # The distillery name must be two words
            distillery = whiskyNameData[0] + "_" + whiskyNameData[1]
            barrelTypeCode = whiskyNameData[3]
            bondYearAndQuarter = whiskyNameData[2]

        else:
            distillery = whiskyNameData[0]
            barrelTypeCode = whiskyNameData[2]
            bondYearAndQuarter = whiskyNameData[1]

        
        minAvgBuyPrice = 100
        maxAvgBuyPrice = 0
        # Have to remove the first cell in the row as it is the title of the row.
        # Need to convert avgBuyCol into a list from a tuple in order to change it.
        avgBuyCol = list(avgBuyCol)
        avgBuyCol.remove(avgBuyCol[0])
        for cell in avgBuyCol:
            if type(cell.value) == float:
                if (cell.value) < minAvgBuyPrice:
                    minAvgBuyPrice = cell.value
                if cell.value > maxAvgBuyPrice:
                    maxAvgBuyPrice = cell.value


        minAvgSellPrice = 100
        maxAvgSellPrice = 0
        avgSellCol = list(avgSellCol)
        avgSellCol.remove(avgSellCol[0])
        for cell in avgSellCol:
            if type(cell.value) != "NoneType":
                if cell.value < minAvgSellPrice:
                    minAvgSellPrice = cell.value
                if cell.value > maxAvgSellPrice:
                    maxAvgSellPrice = cell.value


# MAKE IT SO THAT THE YOUNGEST WHISKY IS AGE 0 AND MAKE THE X AXIS BE AGE OF WHISKY IN DAYS
def normalize_bond_date(whisky_group_object):
    # This takes all the whiskies in a whisky group and pushes the whiskies with older bond dates forward such that
    # the prices are normalized allowing comparison of prices between whiskies of different bond dates.
    df_list_of_tuples = whisky_group_object.get_group_dfs("historic")
    distillery = whisky_group_object.distillery
    barrel_code = whisky_group_object.barrel_code
    # The list is already in order from oldest fill date to most recent
    recent_fill_period = df_list_of_tuples[len(df_list_of_tuples) - 1][0]
    most_recent_whisky = Whisky(distillery, barrel_code, fill_period=recent_fill_period)
    recent_bond_year = int(most_recent_whisky.bond_year)
    recent_bond_quarter_number = int((most_recent_whisky.bond_quarter).replace('Q', ''))  # This changes (E.G) Q4 into just 4

    recent_whisky_df = df_list_of_tuples[len(df_list_of_tuples) - 1][1]
    list_of_normalized_dfs = []
    for tuple_entry in df_list_of_tuples[0:len(df_list_of_tuples) - 5]:  # Exclude the last entry (most recent whisky)
        fill_period = tuple_entry[0]
        df = tuple_entry[1]
        whisky = Whisky(distillery, barrel_code, fill_period=fill_period)
        bond_year = int(whisky.bond_year)
        bond_quarter_number = int((whisky.bond_quarter).replace('Q', ''))  # This changes (E.G) Q4 into just 4
        period_difference_quarters = ((recent_bond_year - bond_year) * 4) + (recent_bond_quarter_number - bond_quarter_number)  # In number of quarters
        period_difference = period_difference_quarters * LENGTH_OF_QUARTER
        period_difference_dt = datetime.timedelta(period_difference)  # A timedelta object
        df[DATE_COLUMN_LABEL] += period_difference_dt
        list_of_normalized_dfs.append(df)

    list_of_normalized_dfs.append(recent_whisky_df)

    final_df = list_of_normalized_dfs[0].set_index("Date")
    for df in list_of_normalized_dfs[1:]:
        # Starting with the second element in the list, since we have already set final_df to be the first element.
        final_df = pd.merge(final_df, df.set_index("Date"), left_index=True, right_index=True, how="outer")
    final_df.fillna(method="ffill", inplace=True)  # When a price/date pair is missing for a column, use the last price

    final_df.to_excel("normalized_bond_dates.xlsx", "sheet1")
    whisky_group = distillery.capitalize() + ' ' + barrel_code.upper()
    final_df.plot(title=whisky_group, figsize=(16, 9))
    plt.show()

group = WhiskyGroup("ardmore", "BBR")
normalize_bond_date(group)
