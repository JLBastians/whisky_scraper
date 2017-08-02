'''
This makes an excel file with all the whisky names, barrel codes and barrel age copied into an excel file.
Does not automatically get the exact barrel date, that needs to be input manually into the excel file.
'''

from openpyxl import Workbook, load_workbook
from whisky_scraper import GetWhiskyList
from os import getcwd, path, listdir
from proforma import METADATA_FILENAME
from WhiskyClass import Whisky
import pandas as pd

'''
FUTURE IMPROVEMENTS:
    Right now, new whiskies are appended to the end of the file. It would be neat if we could order the whiskies once
    the file has been updated.
'''
def update_workbook():
    file = get_file()
    wb = file["workbook"]
    ws = file["worksheet"]
    whisky_list_json = GetWhiskyList("https://www.whiskyinvestdirect.com/cameronbridge/2013/Q3/HHR/chart.do")  # Any URL will work.
    final_row = len(ws["A"])  # Represents he row number of the last entry in the file.
    for whisky in whisky_list_json:
        distillery = whisky["formattedDistillery"]
        barrel_code = whisky["barrelTypeCode"]
        whisky_type = whisky["categoryName"]  # Not a part of the URL, just extra info
        for bondPeriod in whisky["bonds"]:
            bond_year = bondPeriod["bondYear"]
            bond_quarter = bondPeriod["bondQuarter"]
            whisky_obj = Whisky(distillery, barrel_code, bond_year, bond_quarter)  # Instantiating this only to get the whisky_code
            whisky_code = whisky_obj.whisky_code
            if not whisky_entry_exists(whisky_code, ws):
                # The whisky cannot be found in the file, so we need to add it.
                final_row = str(int(final_row) + 1)  # We want to add the new whisky below the final entry
                ws["A" + final_row] = distillery
                ws["B" + final_row] = str(bond_year) + "/" + bond_quarter
                ws["C" + final_row] = barrel_code
                ws["D" + final_row] = whisky_type
                # The start date (in column E) must be set manually.
                ws["F" + final_row] = whisky_code
    check_dates(ws)
    wb.save(filename=METADATA_FILENAME)


def check_dates(worksheet):
    date_col = worksheet['E'][1:]  # Start at 1 since the 0th entry contains the column title.
    row_index = 2  # Starting at row 2 which is the first row below the column title.
    need_update = False
    for date_cell in date_col:
        if date_cell.value is None:
            # The row entry has no start date and so we need to ask for one.
            print(worksheet["A" + str(row_index)].value + ", " + worksheet["B" + str(row_index)].value + ", "
                  + worksheet["C" + str(row_index)].value + " has no start date.")
            print("Please enter the start date manually into cell E" + str(row_index))
            print("File directory: " + path.join(getcwd(), METADATA_FILENAME) + "\n")
            need_update = True
        row_index += 1
    if not need_update:
        print("Everything is up to date.")


def whisky_entry_exists(whisky_code, worksheet):
    code_col = worksheet['F'][1:]
    for code_cell in code_col:
        if whisky_code == code_cell.value:
            # The whisky already exists in the file, no need to continue.
            whisky_already_exists = True
            return whisky_already_exists
    # At this point, we know the whisky isn't in the file, so we need to add it.
    return False


def get_file():
    if METADATA_FILENAME in listdir(getcwd()):
        # The file exists and can be simply loaded.
        wb = load_workbook(METADATA_FILENAME)
        ws = wb.get_active_sheet()
        ws_and_wb = {"worksheet": ws, "workbook": wb}
        return ws_and_wb

    else:
        # The file does not exist (at least not in the working directory), o needs to be created.
        wb = Workbook()
        ws = wb.active  # Select the default worksheet
        ws.title = "Whisky Metadata"

        # Setting up the excel table
        ws["A1"] = "Distiller"
        ws["B1"] = "Cask Filled"
        ws["C1"] = "Barrel Code"
        ws["D1"] = "Whisky Type"
        ws["E1"] = "Date Listed"
        ws["F1"] = "Whisky Code"

        ws_and_wb = {"worksheet": ws, "workbook": wb}
        return ws_and_wb


update_workbook()
